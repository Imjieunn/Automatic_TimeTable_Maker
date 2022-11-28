# 시간표 가져오는 모듈 
from selenium import webdriver
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import Workbook,load_workbook
from selenium.webdriver.common.by import By
from app import uid

# Chrome의 경우 | 아까 받은 chromedriver의 위치를 지정해준다.
chromedrvier = 'C:\\Users\\user\\OSS\\chromedriver.exe'
driver = webdriver.Chrome(chromedrvier)

# url에 접근한다.
driver.get('https://everytime.kr/login')
# 암묵적으로 웹 자원 로드를 위해 5초까지 기다려 준다.
driver.implicitly_wait(5)

# 아이디/비밀번호를 입력해준다. (직접 입력)
driver.find_element(By.NAME, 'userid').send_keys(uid) #send_keys()부분에 직접 아이디 입력
driver.find_element(By.NAME,'password').send_keys("dlawldms7761!") #send_keys()부분에 직접 비번 입력

# 로그인 버튼을 눌러주자.
driver.find_element(By.XPATH, '//*[@id="container"]/form/p[3]/input').click()

# 시간표 접근 
driver.get('https://everytime.kr/timetable')

#수업 목록에서 검색 클릭
driver.find_element(By.XPATH,'//*[@id="container"]/ul/li[1]').click()

#팝업창 닫기
sleep(2)
#driver.find_element(By.XPATH,'//*[@id="subjects"]/div[1]/a[1]').click()

pre_count = 0
#스크롤 맨아래로 내리기
while True:
    #tr요소 접근
    element = driver.find_elements(By.CSS_SELECTOR, "#subjects > div.list > table > tbody > tr")

    # tr 마지막 요소 접근
    result = element[-1]
    #마지막요소에 focus주기
    driver.execute_script('arguments[0].scrollIntoView(true);',result)
    sleep(2)

    #현재 접근한 요소의 갯수
    current_count = len(element)
    if pre_count == current_count:
        break
    #같지않다면
    pre_count = current_count


html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

trs = soup.select('#subjects > div.list > table > tbody > tr')

results = []

for tr in trs:
    result=[]
    tds = tr.select('#subjects > div.list > table > tbody > tr > td')
    result.append(tds[0].text) #과목코드
    result.append(tds[1].text) #과목명
    result.append(tds[2].text) #교수명
    result.append(tds[3].text.split(',')) #강의시간
    result.append(tds[4].text) #강의실
    result.append(tds[5].text) #구분
    result.append(tds[6].text) #학점
    result.append(tds[10].text) #기타
    results.append(result)
#값이 들어있다면!
if results:
    print("성공!!")
excel_column = 9
write_wb = Workbook()
write_ws = write_wb.create_sheet('result_subject_data.xls')
for data in results:
    write_ws = write_wb.active
    write_ws.append(data)
write_wb.save('C:\\Users\\user\\OSS\\result_subject_data.csv')
